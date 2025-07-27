import os
import re
import tkinter as tk
from tkinter import ttk
from collections import defaultdict
import csv
from tkinter import filedialog, simpledialog, messagebox
from openpyxl import load_workbook
from number_parser import parse_ordinal
import pythoncom
import win32com.client as win32

from log_manager import LogManager
from ordnungszahlen import ordinal_map
# ✅ config.py einbinden
try:
    from Eingabe.config import GLOBALORDNER
except ImportError:
    raise ImportError("config.py mit GLOBALORDNER fehlt oder GLOBALORDNER ist nicht definiert.")

TEXT_DIR = GLOBALORDNER['txt']

# Tags zum Entfernen
START_TAGS = {
    "Einrueckung": "|EinrueckungStart| ",
    "Zentriert": "|ZentriertStart| ",
    "Rechtsbuendig": "|RechtsbuendigStart| ",
}
END_TAGS = {
    "Einrueckung": "|EinrueckungEnde| ",
    "Zentriert": "|ZentriertEnde| ",
    "Rechtsbuendig": "|RechtsbuendigEnde| ",
}
IGNORE_TAGS = set(START_TAGS.values()) | set(END_TAGS.values())

# Dateiname: Kapitel_123.txt
FILE_PATTERN = re.compile(r'^(.+?)_(\d{3})\.txt$')

def roman_to_int(s):
    print(f"[DEBUG] roman_to_int called with: {s}")

    # Nur Großbuchstaben und römische Zeichen erlaubt
    if not re.fullmatch(r'[IVXLCDM]+', s.upper()):
        print(f"[DEBUG] Ungültige Zeichen in römischer Zahl: {s}")
        return None

    roman_numerals = {
        'M': 1000, 'CM': 900, 'D': 500, 'CD': 400,
        'C': 100, 'XC': 90, 'L': 50, 'XL': 40,
        'X': 10, 'IX': 9, 'V': 5, 'IV': 4, 'I': 1
    }
    i = 0
    result = 0
    s = s.upper()
    while i < len(s):
        # Prüfe 2-Zeichen Kombination
        if i+1 < len(s) and s[i:i+2] in roman_numerals:
            result += roman_numerals[s[i:i+2]]
            i += 2
        elif s[i] in roman_numerals:
            result += roman_numerals[s[i]]
            i += 1
        else:
            print(f"[DEBUG] Ungültiges römisches Zeichen gefunden: {s[i]}")
            return None  # Ungültig

    print(f"[DEBUG] roman_to_int result: {result}")
    return result


def chapter_sort_key(chapter_name):
    print(f"[DEBUG] chapter_sort_key called with: {chapter_name}")
    prefix = chapter_name.split("_")[0].strip()  # z. B. "IX. Kapitelname" oder "Prolog"

    # 1. Prolog ganz vorne, egal wie geschrieben
    if prefix.lower().startswith("prolog"):
        print("[DEBUG] Kapitel ist Prolog")
        return (0, 0)

    # 2. Römische Zahl mit Punkt am Anfang (z. B. "IX. Die Zeit")
    match = re.match(r'^([IVXLCDM]+)\.', prefix.upper())
    if match:
        roman = match.group(1)
        value = roman_to_int(roman)
        if value is not None:
            print(f"[DEBUG] Römische Zahl erkannt: {roman} -> {value}")
            return (1, value)

    # 3. Alles andere kommt danach
    print("[DEBUG] Kapitel als String sortieren:", prefix)
    return (2, prefix)

def clean_and_count_words(filepath):
    print(f"[DEBUG] clean_and_count_words reading file: {filepath}")
    with open(filepath, "r", encoding="utf-8") as f:
        text = f.read()
    for tag in IGNORE_TAGS:
        text = text.replace(tag, "")
    count = len(re.findall(r'\b\w+\b', text, flags=re.UNICODE))
    print(f"[DEBUG] Wortanzahl in {filepath}: {count}")
    return count

def process_files(directory):
    print(f"[DEBUG] process_files called on directory: {directory}")
    chapter_files = defaultdict(list)
    max_len = 0

    for filename in sorted(os.listdir(directory)):
        print(f"[DEBUG] Datei prüfen: {filename}")
        match = FILE_PATTERN.match(filename)
        if match:
            chapter = match.group(1)
            path = os.path.join(directory, filename)
            count = clean_and_count_words(path)
            chapter_files[chapter].append(count)
            print(f"[DEBUG] {filename} zu Kapitel {chapter} mit Wortanzahl {count} hinzugefügt")

    max_len = max(len(v) for v in chapter_files.values()) if chapter_files else 0
    print(f"[DEBUG] Max. Anzahl von Dateien in einem Kapitel: {max_len}")

    sorted_chapters = sorted(chapter_files.keys(),  key=chapter_sort_key)
    print(f"[DEBUG] Sortierte Kapitel: {sorted_chapters}")

    table_rows = []
    for i in range(max_len):
        row = {}
        for chapter in sorted_chapters:
            if i < len(chapter_files[chapter]):
                row[chapter] = chapter_files[chapter][i]
            else:
                row[chapter] = ""
        table_rows.append(row)

    print(f"[DEBUG] Tabelle mit {len(table_rows)} Zeilen erstellt")
    return table_rows, sorted_chapters, chapter_files

def export_to_txt(table_rows, chapters, chapter_files):
    import os
    export_dir = os.path.dirname(TEXT_DIR)  # eine Ebene höher als 'txt'
    filepath = os.path.join(export_dir, "wortzaehlung_export.txt")
    print(f"[DEBUG] export_to_txt: Exportiere nach {filepath}")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("Nr.\t" + "\t".join(chapters) + "\n")

        for idx, row in enumerate(table_rows, 1):
            zeile = [str(idx)] + [str(row[chap]) for chap in chapters]
            f.write("\t".join(zeile) + "\n")

        summe = ["→ Summe"] + [str(sum(chapter_files[chap])) for chap in chapters]
        f.write("\t".join(summe) + "\n")

    print(f"[DEBUG] TXT Export abgeschlossen")

def export_to_csv(table_rows, chapters, chapter_files):
    export_dir = os.path.dirname(TEXT_DIR)
    filepath = os.path.join(export_dir, "wortzaehlung_export.csv")
    print(f"[DEBUG] export_to_csv: Exportiere nach {filepath}")

    with open(filepath, "w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Nr."] + chapters)

        for idx, row in enumerate(table_rows, 1):
            writer.writerow([idx] + [row[chap] for chap in chapters])

        sum_row = ["→ Summe"] + [sum(chapter_files[chap]) for chap in chapters]
        writer.writerow(sum_row)

    print(f"[DEBUG] CSV Export abgeschlossen")

def open_excel_update_dialog(table_rows, chapters, chapter_files):
    import datetime

    print("[DEBUG] open_excel_update_dialog gestartet")

    dialog = tk.Toplevel()
    dialog.title("Excel-Aktualisierung")
    dialog.geometry("400x300")

    def choose_file():
        path = filedialog.askopenfilename(filetypes=[("Excel-Dateien", "*.xlsx *.xls")])
        excel_path_var.set(path)
        print(f"[DEBUG] Excel-Datei gewählt: {path}")

    default_excel_path = os.path.join(
        os.path.dirname(os.path.dirname(GLOBALORDNER["Eingabe"])), "wortzaehlung_vergleich.xlsx")
    excel_path_var = tk.StringVar(value=default_excel_path)

    tk.Label(dialog, text="Excel-Datei wählen:").pack(pady=5)
    tk.Button(dialog, text="Datei auswählen", command=choose_file).pack()
    tk.Label(dialog, textvariable=excel_path_var, wraplength=380, fg="blue", font=("Arial", 8)).pack()

    tk.Label(dialog, text="Versionsjahr:").pack()
    year_var = tk.StringVar(value=str(datetime.datetime.now().year))
    tk.Entry(dialog, textvariable=year_var).pack()

    tk.Label(dialog, text="Ignoriere Wortzahlen unter:").pack()
    min_word_var = tk.StringVar(value="20")
    combo = ttk.Combobox(dialog, textvariable=min_word_var, values=[10, 15, 20, 25, 30])
    combo.pack()

    tk.Label(dialog, text='Kulissen-Kapitel erkennen (optional):').pack()
    kulisse_var = tk.StringVar(value="Kulissen")
    tk.Entry(dialog, textvariable=kulisse_var).pack()

    def submit():
        print("[DEBUG] submit in open_excel_update_dialog aufgerufen")
        if not excel_path_var.get():
            messagebox.showerror("Fehler", "Bitte wähle eine Excel-Datei aus.")
            print("[DEBUG] Keine Excel-Datei gewählt")
            return
        try:
            year = int(year_var.get())
            min_words = int(min_word_var.get())
            print(f"[DEBUG] Version: {year}, Mindestwortanzahl: {min_words}")
        except ValueError:
            messagebox.showerror("Fehler", "Versionsjahr und Mindestwortanzahl müssen Zahlen sein.")
            print("[DEBUG] Ungültige Eingabe für Jahr oder Mindestwortanzahl")
            return

        dialog.destroy()
        update_excel(
            excel_path_var.get(), table_rows, chapters, chapter_files,
            version_year=year,
            min_word_threshold=min_words,
            kulisse_filter=kulisse_var.get().strip()
        )

    tk.Button(dialog, text="Aktualisieren", command=submit, bg="#cfe2f3").pack(pady=10)

def int_to_roman(n):
    print(f"[DEBUG] int_to_roman called with: {n}")
    val_map = [
        (1000, 'M'), (900, 'CM'), (500, 'D'), (400, 'CD'),
        (100, 'C'),  (90, 'XC'),  (50, 'L'),  (40, 'XL'),
        (10, 'X'),   (9, 'IX'),   (5, 'V'),   (4, 'IV'),
        (1, 'I')
    ]
    result = ""
    for val, sym in val_map:
        while n >= val:
            result += sym
            n -= val
    print(f"[DEBUG] int_to_roman result: {result}")
    return result

def excel_file_opened(path):
    pythoncom.CoInitialize()
    excel = win32.Dispatch("Excel.Application")
    for wb in excel.Workbooks:
        if wb.FullName.lower() == os.path.abspath(path).lower():
            return wb
    return None

def update_excel(excel_path, table_rows, chapters, chapter_files, version_year, min_word_threshold, kulisse_filter):
    print(f"[DEBUG] update_excel gestartet mit Datei: {excel_path}")

    wb_excel = excel_file_opened(excel_path)
    if wb_excel:
        print(f"[DEBUG] Datei '{excel_path}' ist offen in Excel. Speichere und schließe sie.")
        wb_excel.Save()
        wb_excel.Close(False)
        print(f"[DEBUG] Datei '{excel_path}' wurde gespeichert und geschlossen in Excel.")

    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        col_map = {}
        for col in range(2, ws.max_column + 1):
            jahr = str(ws.cell(row=1, column=col).value).strip() if ws.cell(row=1, column=col).value else ""
            kapitelkopf = str(ws.cell(row=2, column=col).value).strip() if ws.cell(row=2, column=col).value else ""
            if jahr == str(version_year) and kapitelkopf:
                col_map[kapitelkopf] = col
        print(f"[DEBUG] Gefundene Spalten für Jahr {version_year}: {col_map}")

        for col_header, col in col_map.items():
            kapitel = next((k for k in chapters if col_header in k), None)
            if not kapitel:
                print(f"[DEBUG] Kapitel {col_header} nicht gefunden in chapters")
                continue
            counts = chapter_files.get(kapitel, [])
            row = 3
            for count in counts:
                if isinstance(count, (int, float)) and count >= min_word_threshold:
                    ws.cell(row=row, column=col, value=count)
                    print(f"[DEBUG] Schreibe {count} in Zeile {row}, Spalte {col}")
                    row += 1

        if kulisse_filter:
            print(f"[DEBUG] Kulisse Filter angewendet: {kulisse_filter}")
            for kapitel, counts in chapter_files.items():
                print(f"[DEBUG] Prüfe Kapitel '{kapitel}' auf Filter '{kulisse_filter}'")
                if kulisse_filter.lower() not in kapitel.lower():
                    continue

                kapitel_lc = kapitel.lower()
                ziel_spaltenheader = kulisse_filter

                if "prolog" in kapitel_lc:
                    ziel_spaltenheader = next(
                        (hdr for hdr in col_map if "prolog" in hdr.lower()),
                        None
                    )

                elif "schlussszene" in kapitel_lc:
                    ziel_spaltenheader = next(
                        (hdr for hdr in col_map if "clausula" in hdr.lower()),
                        None
                    )
                    
                else:
                    match = re.search(r"(des|der)\s+([a-zäöüß]+)", kapitel_lc)
                    if match:
                        wort = match.group(2)
                        print(f"[DEBUG] Ordinal-Wort im Kapitel gefunden: '{wort}'")

                        zahl = ordinal_map.get(wort)
                        print(f"[DEBUG] Ermittelte Zahl für '{wort}': {zahl}")

                        if zahl:
                            roem = int_to_roman(zahl) + "."
                            ziel_spaltenheader = next(
                                (hdr for hdr in col_map if roem in hdr),
                                None
                            )

                if ziel_spaltenheader:
                    ziel_col = col_map.get(ziel_spaltenheader)
                    print(f"[DEBUG] Ziel Spaltenheader: {ziel_spaltenheader}, Spalte: {ziel_col}")

                    ziel_zeile = None
                    for row in range(3, ws.max_row + 1):
                        spalte_a = ws.cell(row=row, column=1).value
                        if spalte_a and kulisse_filter.lower() in str(spalte_a).lower():
                            ziel_zeile = row
                            break

                    if ziel_zeile is None:
                        print(f"[DEBUG] Kein Eintrag mit '{kulisse_filter}' in Spalte A gefunden. Schreibe in erste freie Zeile.")
                        for row in range(3, ws.max_row + 2):
                            if not ws.cell(row=row, column=1).value:
                                ziel_zeile = row
                                break

                    if ziel_zeile and ziel_col:
                        ws.cell(row=ziel_zeile, column=ziel_col, value=sum(counts))
                        print(f"[DEBUG] Kulissenwert {sum(counts)} in Zeile {ziel_zeile}, Spalte {ziel_col} geschrieben")
                    else:
                        print(f"[DEBUG] Konnte keine passende Zeile oder Spalte für '{kulisse_filter}' finden.")

        wb.save(excel_path)
        print("[DEBUG] Excel-Update erfolgreich abgeschlossen")

        if wb_excel:
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = True
            excel.Workbooks.Open(os.path.abspath(excel_path))
            print(f"[DEBUG] Datei '{excel_path}' wurde in Excel wieder geöffnet.")

        messagebox.showinfo("Erfolg", f"Excel-Datei aktualisiert:\n{excel_path}")

    except Exception as e:
        messagebox.showerror("Fehler", f"Fehler beim Excel-Update:\n{str(e)}")
        print(f"[DEBUG] Fehler beim Excel-Update: {e}")


def create_gui(table_rows, chapters, chapter_files):
    print("[DEBUG] create_gui gestartet")
    root = tk.Tk()
    root.title("Wortanzahl pro Datei (Kapitel = Spalten)")

    columns = ["Nr."] + chapters
    tree = ttk.Treeview(root, columns=columns, show="headings")

    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center", width=120)

    for idx, row in enumerate(table_rows, 1):
        values = [idx] + [row[chapter] for chapter in chapters]
        tree.insert("", "end", values=values)

    sum_row = ["→ Summe"]
    for chapter in chapters:
        summe = sum(chapter_files[chapter])
        sum_row.append(summe)
    tree.insert("", "end", values=sum_row, tags=("summe",))

    tree.tag_configure("summe", background="#ffe5cc", font=("Arial", 10, "bold"))
    tree.pack(fill=tk.BOTH, expand=True)

    export_button = tk.Button(
        root,
        text="Als TXT exportieren",
        command=lambda: export_to_txt(table_rows, chapters, chapter_files),
        bg="#d9ead3",
        font=("Arial", 10, "bold")
    )
    export_button.pack(pady=5)

    export_csv_button = tk.Button(
        root,
        text="Als CSV exportieren",
        command=lambda: export_to_csv(table_rows, chapters, chapter_files),
        bg="#f4cccc",
        font=("Arial", 10, "bold")
    )
    export_csv_button.pack(pady=5)

    excel_button = tk.Button(
        root,
        text="Excel-Aktualisierung starten",
        command=lambda: open_excel_update_dialog(table_rows, chapters, chapter_files),
        bg="#c9daf8",
        font=("Arial", 10, "bold")
    )
    excel_button.pack(pady=10)

    root.mainloop()

if __name__ == "__main__":
    print("Logger wird initialisiert")
    logger = LogManager('meinlog_Komplett.log', extra_logfile='meinLog_letzterDurchlauf.log')
    rows, chapters, files = process_files(TEXT_DIR)
    create_gui(rows, chapters, files)
